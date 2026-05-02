import openpyxl
import shutil
import os
from datetime import datetime
import db

TIPOS = ["Monstruos", "Rituales", "Fusion", "Sincronia", "Enlace", "XYZ", "Pendulo", "Magia", "Trampas"]
SINGLE_BLOCK = {"Rituales"}
CARD_HEADERS = ["Cant", "Nombre", "Ubicación", "Ubicación copia", "Cant 2", "Repetidos"]
DECK_HEADERS = ["Cant", "Nombre", "Copia Extra", "Repetidos"]


def _write_card(ws, row, col, card):
    ws.cell(row=row, column=col,     value=card.get("cant") or 0)
    ws.cell(row=row, column=col + 1, value=card.get("nombre"))
    ws.cell(row=row, column=col + 2, value=card.get("ubicacion"))
    ws.cell(row=row, column=col + 3, value=card.get("ubicacion_copia"))
    c2 = card.get("cant2")
    try:
        ws.cell(row=row, column=col + 4, value=int(c2) if c2 is not None else None)
    except (ValueError, TypeError):
        ws.cell(row=row, column=col + 4, value=None)
    ws.cell(row=row, column=col + 5, value=bool(card.get("repetidos", 0)))


def _write_sheet_headers(ws, is_single):
    for i, h in enumerate(CARD_HEADERS):
        ws.cell(row=1, column=i + 1, value=h)
    if not is_single:
        for i, h in enumerate(CARD_HEADERS):
            ws.cell(row=1, column=i + 8, value=h)


def _write_deck_headers(ws, row):
    for block in range(3):
        start = block * 4 + 1
        for i, h in enumerate(DECK_HEADERS):
            ws.cell(row=row, column=start + i, value=h)


def export_excel(path, progress_cb=None):
    conn = db.get_conn()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    total = len(TIPOS) + 1
    for step, tipo in enumerate(TIPOS):
        ws = wb.create_sheet(tipo)
        is_single = tipo in SINGLE_BLOCK
        _write_sheet_headers(ws, is_single)

        cards = [dict(r) for r in conn.execute(
            "SELECT * FROM cartas WHERE tipo=? ORDER BY id", (tipo,)
        ).fetchall()]

        if is_single:
            for i, card in enumerate(cards):
                if not card.get("vacio"):
                    _write_card(ws, i + 2, 1, card)
        else:
            # Reconstruct paired rows (left block + right block)
            for i in range(0, len(cards), 2):
                excel_row = i // 2 + 2
                left = cards[i]
                right = cards[i + 1] if i + 1 < len(cards) else None
                if not left.get("vacio"):
                    _write_card(ws, excel_row, 1, left)
                if right and not right.get("vacio"):
                    _write_card(ws, excel_row, 8, right)

        if progress_cb:
            progress_cb(step + 1, total)

    # Decks sheet
    ws_d = wb.create_sheet("Decks")
    deck_names = [r[0] for r in conn.execute(
        "SELECT DISTINCT deck_nombre FROM decks ORDER BY deck_nombre"
    ).fetchall()]

    current_row = 1
    for group_start in range(0, len(deck_names), 3):
        group = deck_names[group_start:group_start + 3]
        _write_deck_headers(ws_d, current_row)
        current_row += 1

        group_cards = []
        for dn in group:
            group_cards.append([dict(r) for r in conn.execute(
                "SELECT * FROM decks WHERE deck_nombre=? ORDER BY id", (dn,)
            ).fetchall()])
        while len(group_cards) < 3:
            group_cards.append([])

        max_r = max((len(c) for c in group_cards), default=0)
        for i in range(max_r):
            for bi, cards in enumerate(group_cards):
                if i < len(cards):
                    c = cards[i]
                    col = bi * 4 + 1
                    ws_d.cell(row=current_row + i, column=col,     value=c["cant"])
                    ws_d.cell(row=current_row + i, column=col + 1, value=c["nombre"])
                    c2 = c.get("copia_extra")
                    try:
                        ws_d.cell(row=current_row + i, column=col + 2,
                                  value=int(c2) if c2 is not None else None)
                    except (ValueError, TypeError):
                        ws_d.cell(row=current_row + i, column=col + 2, value=None)
                    ws_d.cell(row=current_row + i, column=col + 3,
                              value=bool(c.get("repetidos", 0)))
        current_row += max_r + 1  # blank separator between groups

    if progress_cb:
        progress_cb(total, total)

    conn.close()
    wb.save(path)


def update_excel(original_path):
    """Hace backup y sobreescribe el original."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base, ext = os.path.splitext(original_path)
    backup_path = f"{base}_backup_{ts}{ext}"
    shutil.copy2(original_path, backup_path)
    export_excel(original_path)
    return backup_path
