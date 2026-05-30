import openpyxl
from openpyxl.utils import range_boundaries
from db import get_conn, init_db

TIPOS = ["Monstruos", "Rituales", "Fusion", "Sincronia", "Enlace", "XYZ", "Pendulo", "Magia", "Trampas"]
SINGLE_BLOCK = {"Rituales"}
MAX_CONSECUTIVE_BLANKS = 30


def _get_table_bounds(path):
    """Devuelve {sheet_name: (data_min_row, max_row)} leyendo los rangos de tabla."""
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    bounds = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        for tbl in ws.tables.values():
            _, min_row, _, max_row = range_boundaries(tbl.ref)
            bounds[sn] = (min_row + 1, max_row)  # +1 salta la cabecera
            break
    wb.close()
    return bounds


def _clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _clean_int(v):
    """Devuelve entero o None; descarta texto no numérico."""
    if v is None:
        return None
    try:
        return int(float(str(v).strip()))
    except (ValueError, TypeError):
        return None


def _to_bool(v):
    if v is None:
        return 0
    return 1 if str(v).strip().lower() in ("true", "1", "verdadero") else 0


def _safe(row, idx, default=None):
    try:
        v = row[idx]
        return v if v is not None else default
    except (IndexError, TypeError):
        return default


def _row_has_data(row):
    return any(
        v is not None and str(v).strip() not in ("", "None")
        for v in (row or [])
    )


def _trim_trailing_vacios(conn, tipo):
    """Elimina huecos vacíos del final (después de la última carta real) de un tipo."""
    rows = conn.execute(
        "SELECT id, vacio FROM cartas WHERE tipo=? ORDER BY id DESC", (tipo,)
    ).fetchall()
    ids_to_delete = []
    for row in rows:
        if row[1]:  # vacio=1
            ids_to_delete.append(row[0])
        else:
            break  # encontrada la última carta real, parar
    if ids_to_delete:
        conn.execute(
            f"DELETE FROM cartas WHERE id IN ({','.join('?' * len(ids_to_delete))})",
            ids_to_delete,
        )


def import_excel(path, progress_cb=None):
    init_db()
    table_bounds = _get_table_bounds(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    conn = get_conn()

    conn.execute("DELETE FROM cartas")
    conn.execute("DELETE FROM decks")

    total_steps = len(TIPOS) + 1  # +1 for decks

    sort_counter = 0

    for step, tipo in enumerate(TIPOS):
        if tipo not in wb.sheetnames:
            if progress_cb:
                progress_cb(step + 1, total_steps)
            continue

        ws = wb[tipo]
        is_single = tipo in SINGLE_BLOCK

        bounds = table_bounds.get(tipo)
        min_row = bounds[0] if bounds else 2

        all_rows = list(ws.iter_rows(min_row=min_row, values_only=True))

        # Determinar rango real de cada bloque por separado
        first_l = last_l = first_r = last_r = -1
        for i, row in enumerate(all_rows):
            nl = _clean(_safe(row, 1))
            nr = None if is_single else _clean(_safe(row, 8))
            if nl:
                if first_l == -1:
                    first_l = i
                last_l = i
            if nr:
                if first_r == -1:
                    first_r = i
                last_r = i

        if first_l == -1 and first_r == -1:
            if progress_cb:
                progress_cb(step + 1, total_steps)
            continue

        last_idx = max(last_l, last_r)
        all_rows = all_rows[:last_idx + 1]

        for i, row in enumerate(all_rows):
            l_active = first_l != -1 and first_l <= i <= last_l
            r_active = not is_single and first_r != -1 and first_r <= i <= last_r
            nombre_l = _clean(_safe(row, 1)) if l_active else None
            nombre_r = _clean(_safe(row, 8)) if r_active else None

            if nombre_l == "Nombre":
                nombre_l = None
            if nombre_r == "Nombre":
                nombre_r = None

            left_has = nombre_l is not None
            right_has = nombre_r is not None

            if l_active:
                conn.execute(
                    "INSERT INTO cartas (tipo,cant,nombre,ubicacion,ubicacion_copia,cant2,repetidos,vacio,sort_order) "
                    "VALUES (?,?,?,?,?,?,?,?,?)",
                    (
                        tipo,
                        _safe(row, 0, 0) or 0,
                        nombre_l,
                        _clean(_safe(row, 2)),
                        _clean(_safe(row, 3)),
                        _safe(row, 4),
                        _to_bool(_safe(row, 5)),
                        0 if left_has else 1,
                        sort_counter,
                    ),
                )
                sort_counter += 1

            if r_active:
                conn.execute(
                    "INSERT INTO cartas (tipo,cant,nombre,ubicacion,ubicacion_copia,cant2,repetidos,vacio,sort_order) "
                    "VALUES (?,?,?,?,?,?,?,?,?)",
                    (
                        tipo,
                        _safe(row, 7, 0) or 0,
                        nombre_r,
                        _clean(_safe(row, 9)),
                        _clean(_safe(row, 10)),
                        _safe(row, 11),
                        _to_bool(_safe(row, 12)),
                        0 if right_has else 1,
                        sort_counter,
                    ),
                )
                sort_counter += 1

        # Eliminar huecos vacíos al final (trailing) de este tipo
        _trim_trailing_vacios(conn, tipo)

        if progress_cb:
            progress_cb(step + 1, total_steps)

    # Decks sheet — cada cabecera (Nombre/Cant) marca inicio de nueva sección de decks
    if "Decks" in wb.sheetnames:
        ws = wb["Decks"]
        deck_section = 0  # se incrementa con cada fila cabecera
        for row in ws.iter_rows(min_row=1, values_only=True):
            if not row:
                continue
            # Detectar fila cabecera: segunda celda es "Nombre"
            if _clean(_safe(row, 1)) == "Nombre":
                deck_section += 1
                continue
            for block_idx, start in enumerate((0, 4, 8)):
                nombre = _clean(_safe(row, start + 1))
                if nombre:
                    deck_nombre = f"Deck {(deck_section - 1) * 3 + block_idx + 1}"
                    conn.execute(
                        "INSERT INTO decks (cant,nombre,copia_extra,repetidos,deck_nombre) "
                        "VALUES (?,?,?,?,?)",
                        (
                            _safe(row, start, 1) or 1,
                            nombre,
                            _clean_int(_safe(row, start + 2)),
                            _to_bool(_safe(row, start + 3)),
                            deck_nombre,
                        ),
                    )

    # Estadistica: total from O26 (col 15, row 26)
    if "Estadistica" in wb.sheetnames:
        ws_stat = wb["Estadistica"]
        total_val = ws_stat["O26"].value
        if total_val:
            conn.execute(
                "INSERT OR REPLACE INTO config VALUES ('total_cartas',?)",
                (str(int(total_val)),),
            )

    if progress_cb:
        progress_cb(total_steps, total_steps)

    conn.commit()
    conn.close()
    wb.close()
